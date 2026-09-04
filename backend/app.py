from __future__ import annotations

import base64
import binascii
import csv
import gc
import hashlib
import hmac
import io
import json
import logging
import os
import secrets
import threading
import time
import uuid
from concurrent.futures import ThreadPoolExecutor
from contextlib import asynccontextmanager
from datetime import date, datetime
from pathlib import Path
from urllib.parse import quote, urlsplit

from dotenv import load_dotenv
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field

from src.portfolio.ai import (
    AIError,
    AIResponseError,
    DEFAULT_TOPIC_CLASSIFICATION,
    analyze_frames,
    classify_topics,
    validate_ollama_model,
)
from src.portfolio.database import Database, utc_now
from src.portfolio.ingestion import (
    SpreadsheetValidationError,
    read_spreadsheet,
)
from src.portfolio.jw_session import (
    JWBrowserSession,
    JWSessionError,
)
from src.portfolio.jwplayer import (
    JWPlayerClient,
)
from src.portfolio.frames import extract_frames
from src.portfolio.transcription import transcribe_hls


# ==========================================================
# CONFIGURAÇÃO
# ==========================================================

BASE_DIR = Path(__file__).resolve().parent

load_dotenv(
    BASE_DIR / ".env",
    override=False,
)

# Não substitui uma configuração de logging já existente no
# processo (ex.: uvicorn) — é um no-op se o root logger já
# tiver handlers configurados.
logging.basicConfig(
    level=logging.INFO,
    format=(
        "%(asctime)s "
        "%(levelname)s "
        "%(name)s: "
        "%(message)s"
    ),
)

logger = logging.getLogger(
    "cetrus.portfolio"
)

DATA_DIR = Path(
    os.getenv(
        "CETRUS_DATA_DIR",
        BASE_DIR / "data",
    )
)

DATABASE = Database(
    DATA_DIR / "portfolio.db"
)

WEB_DIR = BASE_DIR / "web"


# ==========================================================
# ACESSO DO FRONTEND
# ==========================================================

def env_flag(
    name: str,
    default: str = "false",
) -> bool:

    return str(
        os.getenv(name, default)
    ).strip().casefold() in {
        "1",
        "true",
        "yes",
        "on",
    }


def configured_origins() -> tuple[str, ...]:

    origins = tuple(
        origin.strip().rstrip("/")
        for origin in os.getenv(
            "CORS_ALLOWED_ORIGINS",
            "",
        ).split(",")
        if origin.strip()
    )

    if "*" in origins:
        raise RuntimeError(
            "CORS_ALLOWED_ORIGINS não pode usar '*' em requisições "
            "autenticadas."
        )

    return origins


CORS_ALLOWED_ORIGINS = configured_origins()

AUTH_ENABLED = env_flag("APP_AUTH_ENABLED")

AUTH_USERNAME = os.getenv(
    "APP_AUTH_USERNAME",
    "",
).strip()

AUTH_PASSWORD = os.getenv(
    "APP_AUTH_PASSWORD",
    "",
)

AUTH_SESSION_SECRET = os.getenv(
    "APP_AUTH_SESSION_SECRET",
    "",
)

AUTH_SESSION_TTL_SECONDS = int(
    os.getenv(
        "APP_AUTH_SESSION_TTL_SECONDS",
        "28800",
    )
)

AUTH_LOGIN_MAX_ATTEMPTS = int(
    os.getenv(
        "APP_AUTH_LOGIN_MAX_ATTEMPTS",
        "5",
    )
)

AUTH_LOGIN_WINDOW_SECONDS = int(
    os.getenv(
        "APP_AUTH_LOGIN_WINDOW_SECONDS",
        "300",
    )
)

AUTH_TRUST_PROXY_HEADERS = env_flag(
    "APP_AUTH_TRUST_PROXY_HEADERS",
)

AUTH_SESSION_VERSION = secrets.token_urlsafe(32)

LOGIN_FAILURES: dict[str, list[float]] = {}
LOGIN_FAILURES_LOCK = threading.Lock()
MAX_LOGIN_SOURCES = 1000

if AUTH_ENABLED:
    if not AUTH_USERNAME or not AUTH_PASSWORD:
        raise RuntimeError(
            "APP_AUTH_USERNAME e APP_AUTH_PASSWORD são obrigatórios "
            "quando APP_AUTH_ENABLED=true."
        )

    if len(AUTH_SESSION_SECRET) < 32:
        raise RuntimeError(
            "APP_AUTH_SESSION_SECRET deve ter pelo menos 32 caracteres."
        )

    if AUTH_SESSION_TTL_SECONDS <= 0:
        raise RuntimeError(
            "APP_AUTH_SESSION_TTL_SECONDS deve ser maior que zero."
        )

    if (
        AUTH_LOGIN_MAX_ATTEMPTS <= 0
        or AUTH_LOGIN_WINDOW_SECONDS <= 0
    ):
        raise RuntimeError(
            "Os limites de tentativas de login devem ser maiores que zero."
        )

# ==========================================================
# BIBLIOTECAS JW PLAYER
# ==========================================================

JW_LIBRARIES = {

    "DEV": {
        "name": "DEV",
        "property_id": "FvJr6FNj",
        "url": "https://dashboard.jwplayer.com/p/FvJr6FNj/media",
    },

    "EBSERH": {
        "name": "EBSERH",
        "property_id": "UBP82vRQ",
        "url": "https://dashboard.jwplayer.com/p/UBP82vRQ/media",
    },

    "SANARFLIX": {
        "name": "SANARFLIX",
        "property_id": "XK8A5jD7",
        "url": "https://dashboard.jwplayer.com/p/XK8A5jD7/media",
    },

    "VIDEOSSANAR": {
        "name": "VIDEOS SANAR",
        "property_id": "XdfUPSCL",
        "url": "https://dashboard.jwplayer.com/p/XdfUPSCL/media",
    },
}


DEFAULT_JW_LIBRARY = "VIDEOSSANAR"


# Mantido para compatibilidade com partes antigas do projeto.

JW_PROPERTY_ID = JW_LIBRARIES[
    DEFAULT_JW_LIBRARY
]["property_id"]

JW_LIBRARY_URL = JW_LIBRARIES[
    DEFAULT_JW_LIBRARY
]["url"]


# ==========================================================
# CHAVES DAS IAs
# ==========================================================

AI_API_KEYS = {

    "Gemini":
        os.getenv(
            "GEMINI_API_KEY",
            "",
        ).strip(),

    "Claude":
        os.getenv(
            "ANTHROPIC_API_KEY",
            "",
        ).strip(),
}


# ==========================================================
# OLLAMA (SOMENTE LOCAL)
# ==========================================================
#
# Ollama exige um servidor rodando na própria máquina e não
# faz parte da infraestrutura hospedada (Render). Em produção,
# ENABLE_OLLAMA=false remove essa opção sem remover o código —
# continua disponível normalmente em ambiente local.

ENABLE_OLLAMA = str(
    os.getenv(
        "ENABLE_OLLAMA",
        "true",
    )
).strip().casefold() not in {
    "false",
    "0",
    "no",
}

AVAILABLE_PROVIDERS = (

    {
        "Gemini",
        "Claude",
        "Ollama",
    }

    if ENABLE_OLLAMA

    else {
        "Gemini",
        "Claude",
    }
)

AVAILABLE_PROVIDERS_MESSAGE = (
    "Selecione Gemini, Claude ou Ollama."
    if ENABLE_OLLAMA
    else "Selecione Gemini ou Claude."
)


# ==========================================================
# FILTRO DE DATA DE PUBLICAÇÃO (JW PLAYER)
# ==========================================================
#
# Reaproveita a variável já documentada em .env.example
# (antes usada só pela CLI legada). Necessária apenas para
# mídias protegidas na API de delivery do JW Player — a
# maioria das consultas funciona sem token.

JW_DELIVERY_TOKEN = os.getenv(
    "JW_DELIVERY_TOKEN",
    "",
).strip()

PUBLISH_DATE_CHECK_WORKERS = 8


# ==========================================================
# INSTÂNCIAS
# ==========================================================

JW_SESSION = JWBrowserSession()

PROCESSOR = ThreadPoolExecutor(
    max_workers=1,
    thread_name_prefix="video-processing",
)

ANALYSIS_LOCK = threading.Lock()

JOBS: dict[str, dict] = {}

# JOBS nunca era podado individualmente (só era zerado por
# inteiro quando uma nova planilha substituía um lote ativo).
# Em uso real — análises individuais e /api/start-eligible, que
# não passam por essa limpeza — o dicionário crescia sem limite
# pela vida inteira do processo. Isso mantém só os mais
# recentes, o suficiente para a Etapa 3/`/api/jobs` (que já só
# exibe os últimos 100).
MAX_JOBS_HISTORY = 200

JOBS_LOCK = threading.Lock()

ACTIVE_BATCH_LOCK = threading.Lock()

ACTIVE_BATCH: threading.Event | None = None

CURRENT_JW_LIBRARY = DEFAULT_JW_LIBRARY

CURRENT_JW_PROPERTY_ID = JW_LIBRARIES[
    DEFAULT_JW_LIBRARY
]["property_id"]


# ==========================================================
# EXCEÇÕES
# ==========================================================

class BatchCancelled(Exception):
    pass


# ==========================================================
# LIFESPAN
# ==========================================================

@asynccontextmanager
async def lifespan(_: FastAPI):

    yield

    JW_SESSION.close()

    PROCESSOR.shutdown(
        wait=False,
        cancel_futures=True,
    )


# ==========================================================
# APLICAÇÃO
# ==========================================================

app = FastAPI(
    title="Portfólio de vídeos Cetrus",
    version="2.2.0",
    lifespan=lifespan,
)


app.mount(
    "/assets",
    StaticFiles(
        directory=WEB_DIR,
    ),
    name="assets",
)


def encode_session(
    username: str,
) -> str:

    payload = json.dumps(
        {
            "username": username,
            "expires_at": int(time.time())
            + AUTH_SESSION_TTL_SECONDS,
            "version": AUTH_SESSION_VERSION,
        },
        separators=(",", ":"),
        sort_keys=True,
    ).encode("utf-8")

    encoded_payload = base64.urlsafe_b64encode(
        payload
    ).rstrip(b"=")

    signature = hmac.new(
        AUTH_SESSION_SECRET.encode("utf-8"),
        encoded_payload,
        hashlib.sha256,
    ).digest()

    encoded_signature = base64.urlsafe_b64encode(
        signature
    ).rstrip(b"=")

    return (
        encoded_payload.decode("ascii")
        + "."
        + encoded_signature.decode("ascii")
    )


def decode_session(
    token: str,
) -> str | None:

    try:
        encoded_payload, encoded_signature = token.split(
            ".",
            maxsplit=1,
        )

        expected_signature = hmac.new(
            AUTH_SESSION_SECRET.encode("utf-8"),
            encoded_payload.encode("ascii"),
            hashlib.sha256,
        ).digest()

        supplied_signature = base64.urlsafe_b64decode(
            encoded_signature
            + "=" * (-len(encoded_signature) % 4)
        )

        if not hmac.compare_digest(
            supplied_signature,
            expected_signature,
        ):
            return None

        payload = json.loads(
            base64.urlsafe_b64decode(
                encoded_payload
                + "=" * (-len(encoded_payload) % 4)
            )
        )

        if not isinstance(payload, dict):
            return None

        username = str(
            payload.get("username") or ""
        )
        expires_at = int(
            payload.get("expires_at") or 0
        )
        session_version = str(
            payload.get("version") or ""
        )

        if (
            not username
            or username != AUTH_USERNAME
            or expires_at <= int(time.time())
            or not secrets.compare_digest(
                session_version,
                AUTH_SESSION_VERSION,
            )
        ):
            return None

        return username

    except (
        binascii.Error,
        UnicodeDecodeError,
        ValueError,
        TypeError,
        json.JSONDecodeError,
    ):
        return None


def authenticated_username(
    request: Request,
) -> str | None:

    if not AUTH_ENABLED:
        return AUTH_USERNAME or "local"

    parts = request.headers.get(
        "authorization",
        "",
    ).split()

    if (
        len(parts) != 2
        or parts[0].casefold() != "bearer"
        or not parts[1]
    ):
        return None

    return decode_session(parts[1])


def origin_is_allowed(
    request: Request,
) -> bool:

    origin = str(
        request.headers.get("origin") or ""
    ).strip().rstrip("/")

    if not origin:
        return not AUTH_ENABLED

    if origin in CORS_ALLOWED_ORIGINS:
        return True

    origin_parts = urlsplit(origin)
    request_host = str(
        request.headers.get("host") or ""
    ).casefold()

    return (
        origin_parts.scheme in {"http", "https"}
        and origin_parts.netloc.casefold() == request_host
    )


def login_source(
    request: Request,
) -> str:

    if AUTH_TRUST_PROXY_HEADERS:
        forwarded = (
            request.headers.get("cf-connecting-ip")
            or request.headers.get("x-forwarded-for", "").split(",")[0]
        ).strip()

        if forwarded:
            return forwarded

    if request.client:
        return request.client.host

    return "unknown"


def login_is_rate_limited(
    request: Request,
) -> bool:

    key = login_source(request)
    cutoff = time.monotonic() - AUTH_LOGIN_WINDOW_SECONDS

    with LOGIN_FAILURES_LOCK:
        attempts = [
            attempted_at
            for attempted_at in LOGIN_FAILURES.get(key, [])
            if attempted_at > cutoff
        ]
        LOGIN_FAILURES[key] = attempts

        return len(attempts) >= AUTH_LOGIN_MAX_ATTEMPTS


def record_login_failure(
    request: Request,
) -> None:

    key = login_source(request)

    with LOGIN_FAILURES_LOCK:
        if (
            key not in LOGIN_FAILURES
            and len(LOGIN_FAILURES) >= MAX_LOGIN_SOURCES
        ):
            LOGIN_FAILURES.pop(
                next(iter(LOGIN_FAILURES))
            )

        LOGIN_FAILURES.setdefault(key, []).append(
            time.monotonic()
        )


def clear_login_failures(
    request: Request,
) -> None:

    key = login_source(request)

    with LOGIN_FAILURES_LOCK:
        LOGIN_FAILURES.pop(key, None)


def add_cors_middleware(
    application: FastAPI,
    origins: tuple[str, ...],
) -> None:

    application.add_middleware(
        CORSMiddleware,
        allow_origins=list(origins),
        allow_credentials=False,
        allow_methods=[
            "GET",
            "POST",
            "OPTIONS",
        ],
        allow_headers=[
            "Accept",
            "Authorization",
            "Content-Type",
        ],
        expose_headers=[
            "Content-Disposition",
        ],
    )


# StaticFiles não define Cache-Control por padrão. Sem isso, o
# navegador pode reaproveitar app.js/style.css de uma visita
# anterior mesmo depois de o arquivo mudar no servidor — a causa
# real de correções em app.js "não aparecerem" no navegador do
# usuário. no-cache força revalidação (ETag/Last-Modified) a
# cada carregamento, sem exigir cache-busting manual na URL.
@app.middleware("http")
async def _no_cache_for_assets(request, call_next):

    response = await call_next(request)

    if request.url.path.startswith("/assets/"):
        response.headers["Cache-Control"] = "no-cache"

    return response


PUBLIC_API_PATHS = {
    "/api/auth/login",
    "/api/auth/session",
}


@app.middleware("http")
async def _protect_api(request: Request, call_next):

    path = request.url.path

    if (
        path.startswith("/api/")
        and request.method not in {"GET", "HEAD", "OPTIONS"}
        and not origin_is_allowed(request)
    ):
        return JSONResponse(
            status_code=403,
            content={
                "detail": "Origem não autorizada."
            },
        )

    protected_path = (
        path.startswith("/api/")
        and path not in PUBLIC_API_PATHS
    ) or path in {
        "/docs",
        "/openapi.json",
        "/redoc",
    }

    if (
        AUTH_ENABLED
        and protected_path
        and not authenticated_username(request)
    ):
        return JSONResponse(
            status_code=401,
            content={
                "detail": "Autenticação necessária."
            },
        )

    return await call_next(request)


add_cors_middleware(
    app,
    CORS_ALLOWED_ORIGINS,
)


# ==========================================================
# MODELOS
# ==========================================================

class AppLoginRequest(BaseModel):

    username: str = Field(
        min_length=1,
        max_length=200,
    )

    password: str = Field(
        min_length=1,
        max_length=1000,
    )


class LoginRequest(BaseModel):

    library: str = DEFAULT_JW_LIBRARY

    property_id: str = ""

    email: str = ""

    password: str = ""


class SwitchLibraryRequest(BaseModel):

    library: str = DEFAULT_JW_LIBRARY

    property_id: str = ""


class ProcessRequest(BaseModel):

    media_ids: list[str] = Field(
        min_length=1,
        max_length=1000,
    )

    provider: str = "Gemini"

    # Uso exclusivamente interno.
    api_key: str = ""

    model: str = "gemini-flash-latest"

    ollama_url: str = (
        "http://127.0.0.1:11434"
    )

    whisper_model: str = "small"

    analysis_mode: str = "frames"

    frame_count: int = Field(
        default=8,
        ge=4,
        le=16,
    )


class ValidationRequest(BaseModel):

    jwplayer_id: str

    final_category: str

    summary: str = Field(
        max_length=500,
    )

    validated: bool = True


class AnalyzeJWPlayerRequest(BaseModel):

    jwplayer_id: str

    library: str = DEFAULT_JW_LIBRARY

    property_id: str = ""

    provider: str = "Gemini"

    model: str = "gemini-flash-latest"

    ollama_url: str = (
        "http://127.0.0.1:11434"
    )

    whisper_model: str = "small"

    analysis_mode: str = "frames"

    frame_count: int = Field(
        default=8,
        ge=4,
        le=16,
    )

    min_publish_date: str = ""

    include_missing_date: bool = False


class StartEligibleRequest(BaseModel):

    run_id: int = Field(
        ge=1,
    )

    provider: str = "Gemini"

    model: str = "gemini-flash-latest"

    ollama_url: str = (
        "http://127.0.0.1:11434"
    )

    whisper_model: str = "small"

    analysis_mode: str = "frames"

    frame_count: int = Field(
        default=8,
        ge=4,
        le=16,
    )


class BackfillClassificationRequest(BaseModel):

    provider: str = "Gemini"

    model: str = "gemini-flash-latest"

    ollama_url: str = (
        "http://127.0.0.1:11434"
    )


# ==========================================================
# UTILITÁRIOS JW PLAYER
# ==========================================================

def normalize_library(
    library: str | None,
) -> str:

    key = str(
        library or ""
    ).strip().upper()

    if key not in JW_LIBRARIES:

        raise HTTPException(
            status_code=422,
            detail=(
                "Biblioteca JW Player inválida. "
                "Use DEV, EBSERH, SANARFLIX ou VIDEOSSANAR."
            ),
        )

    return key


def get_library_config(
    library: str | None,
) -> dict:

    key = normalize_library(
        library
    )

    return JW_LIBRARIES[key]


def get_current_library() -> dict:

    return JW_LIBRARIES[
        CURRENT_JW_LIBRARY
    ]


def library_from_property_id(
    property_id: str | None,
) -> str | None:

    value = str(
        property_id or ""
    ).strip()

    if not value:
        return None

    for key, config in JW_LIBRARIES.items():

        if config["property_id"] == value:

            return key

    return None


def update_current_library(
    library: str,
) -> dict:

    global CURRENT_JW_LIBRARY
    global CURRENT_JW_PROPERTY_ID

    config = get_library_config(
        library
    )

    CURRENT_JW_LIBRARY = library

    CURRENT_JW_PROPERTY_ID = config[
        "property_id"
    ]

    return config


# ==========================================================
# LINK DA AULA (JW PLAYER)
# ==========================================================
#
# Mesma regra usada pelo frontend em buildJWPlayerMediaUrl()
# (web/app.js) — reproduzida aqui para que a exportação
# CSV/XLSX aponte exatamente para a mesma URL exibida no site.
# Não existe uma segunda lógica: qualquer mudança nesse formato
# precisa ser replicada nos dois lugares.

def build_jwplayer_media_url(
    jwplayer_id: str,
    property_id: str,
) -> str:

    jwplayer_id = str(
        jwplayer_id or ""
    ).strip()

    property_id = str(
        property_id or ""
    ).strip()

    if not jwplayer_id or not property_id:
        return ""

    return (
        f"https://dashboard.jwplayer.com/p/{property_id}"
        f"/media/{quote(jwplayer_id, safe='')}"
    )


# ==========================================================
# CHAVE DA IA
# ==========================================================

def get_provider_api_key(
    provider: str,
) -> str:

    provider = str(
        provider or ""
    ).strip()

    if provider == "Ollama":
        return ""

    if provider not in {
        "Gemini",
        "Claude",
    }:

        raise HTTPException(
            status_code=422,
            detail=(
                "Selecione Gemini, Claude ou Ollama."
            ),
        )

    key = AI_API_KEYS.get(
        provider,
        "",
    )

    if not key:

        env_name = (
            f"{provider.upper()}_API_KEY"
        )

        raise HTTPException(
            status_code=422,
            detail=(
                f"A chave da {provider} "
                f"não está configurada no servidor. "
                f"Configure a variável de ambiente "
                f"{env_name}."
            ),
        )

    return key


# ==========================================================
# CONSTRUIR REQUEST
# ==========================================================

def build_process_request(
    *,
    media_ids: list[str],
    provider: str,
    model: str,
    ollama_url: str,
    whisper_model: str,
    analysis_mode: str,
    frame_count: int,
) -> ProcessRequest:

    provider = str(
        provider or ""
    ).strip()

    if provider not in AVAILABLE_PROVIDERS:

        raise HTTPException(
            status_code=422,
            detail=AVAILABLE_PROVIDERS_MESSAGE,
        )

    if not model.strip():

        raise HTTPException(
            status_code=422,
            detail=(
                "Informe o modelo da IA."
            ),
        )

    return ProcessRequest(

        media_ids=media_ids,

        provider=provider,

        api_key=get_provider_api_key(
            provider
        ),

        model=model.strip(),

        ollama_url=ollama_url.strip(),

        whisper_model=whisper_model.strip(),

        analysis_mode=analysis_mode.strip(),

        frame_count=frame_count,
    )


# ==========================================================
# FILTRO DE DATA DE PUBLICAÇÃO (JW PLAYER)
# ==========================================================
#
# Nova etapa ANTES do processamento/IA: consulta o Publish
# date de cada vídeo na API de delivery do JW Player (já usada
# pelo projeto em src/portfolio/jwplayer.py, só que até agora
# apenas pela CLI legada) e decide se o vídeo é elegível para
# análise, sem baixar vídeo, extrair frame ou chamar IA.
#
# Reaproveita a mesma tabela `videos` (publish_date,
# filter_status, filter_reason, eligible_for_analysis) — não
# apaga nem substitui nenhum registro original da planilha.

def parse_publish_date_cutoff(
    value: str,
) -> date | None:

    value = str(
        value or ""
    ).strip()

    if not value:
        return None

    try:

        return date.fromisoformat(
            value
        )

    except ValueError as exc:

        raise HTTPException(
            status_code=422,
            detail=(
                "Data mínima de publicação inválida. "
                "Utilize o formato AAAA-MM-DD."
            ),
        ) from exc


def classify_publish_date(
    publish_date,
    cutoff: date,
    include_missing_date: bool,
) -> tuple[str, str, bool]:

    """
    Retorna (filter_status, filter_reason, eligible_for_analysis).

    filter_status é a categoria para relatório (eligible /
    filtered / no_date); eligible_for_analysis é o valor que
    realmente controla se o vídeo entra na fila.
    """

    if publish_date is None:

        eligible = bool(
            include_missing_date
        )

        return (
            "no_date",
            (
                "missing_date_included"
                if eligible
                else "date_not_found"
            ),
            eligible,
        )

    if publish_date.date() >= cutoff:

        return (
            "eligible",
            "within_date_range",
            True,
        )

    return (
        "filtered",
        "published_before_cutoff",
        False,
    )


def fetch_publish_date_from_dashboard(
    jwplayer_id: str,
):

    """
    Fallback do publish_date: lê o campo "Created" da página de
    detalhes do vídeo no JW Player (via sessão autenticada), só
    usado quando o playback.json não trouxe pubdate. Nunca
    levanta exceção — retorna None se a sessão não estiver
    conectada ou a data não puder ser lida.
    """

    try:
        return JW_SESSION.fetch_created_date(
            jwplayer_id
        )
    except Exception:
        return None


def check_publish_dates(
    jwplayer_ids: list[str],
    site_id: str,
    min_publish_date: str,
    include_missing_date: bool = False,
) -> dict:

    """
    Consulta o Publish date de cada JWPlayer ID e persiste o
    resultado do filtro em `videos`. Um erro individual não
    interrompe os demais vídeos.

    O Publish date é sempre buscado no JW Player, com ou sem
    `min_publish_date` configurado — esse campo é usado pelo
    filtro de ano da tela de resultados e pela exportação, não
    só pela elegibilidade de processamento. Sem
    `min_publish_date`, a busca acontece do mesmo jeito, mas
    nenhum vídeo é bloqueado por causa da data: todos ficam
    elegíveis (o filtro por data continua opcional).
    """

    summary = {

        "total": len(jwplayer_ids),

        "eligible": 0,

        "filtered": 0,

        "no_date": 0,

        "errors": 0,

        "will_be_analyzed": 0,
    }

    if not jwplayer_ids:
        return summary

    cutoff = parse_publish_date_cutoff(
        min_publish_date
    )

    client = JWPlayerClient(
        site_id=site_id,
        token=JW_DELIVERY_TOKEN,
    )

    def fetch_one(jwplayer_id: str):

        try:

            asset = client.playback(
                jwplayer_id
            )

            publish_date = asset.publish_date

            # Fonte principal (playback.json/pubdate) sem data:
            # tenta a página de detalhes do vídeo no JW Player
            # como alternativa, antes de desistir.
            if publish_date is None:

                publish_date = (
                    fetch_publish_date_from_dashboard(
                        jwplayer_id
                    )
                )

            return (
                jwplayer_id,
                publish_date,
                None,
            )

        except Exception as exc:

            return (
                jwplayer_id,
                None,
                str(exc),
            )

    with ThreadPoolExecutor(
        max_workers=PUBLISH_DATE_CHECK_WORKERS,
    ) as executor:

        results = list(
            executor.map(
                fetch_one,
                jwplayer_ids,
            )
        )

    for jwplayer_id, publish_date, error in results:

        if error is not None:

            logger.warning(
                "Falha ao consultar Publish date "
                "| jwplayer_id=%s erro=%s",
                jwplayer_id,
                error,
            )

            DATABASE.update_filter_result(
                jwplayer_id,
                publish_date=None,
                filter_status="error",
                filter_reason="publish_date_unavailable",
                # Sem filtro configurado, um erro ao buscar a
                # data não deve bloquear o vídeo — só quando há
                # um min_publish_date de fato para respeitar.
                eligible_for_analysis=(cutoff is None),
            )

            summary["errors"] += 1

            if cutoff is None:
                summary["eligible"] += 1
                summary["will_be_analyzed"] += 1

            continue

        if cutoff is None:

            DATABASE.update_filter_result(
                jwplayer_id,
                publish_date=(
                    publish_date.isoformat()
                    if publish_date
                    else None
                ),
                filter_status="eligible",
                filter_reason="no_filter_applied",
                eligible_for_analysis=True,
            )

            summary["eligible"] += 1
            summary["will_be_analyzed"] += 1

            continue

        (
            filter_status,
            filter_reason,
            eligible,
        ) = classify_publish_date(
            publish_date,
            cutoff,
            include_missing_date,
        )

        DATABASE.update_filter_result(
            jwplayer_id,
            publish_date=(
                publish_date.isoformat()
                if publish_date
                else None
            ),
            filter_status=filter_status,
            filter_reason=filter_reason,
            eligible_for_analysis=eligible,
        )

        if filter_status == "eligible":
            summary["eligible"] += 1
        elif filter_status == "no_date":
            summary["no_date"] += 1
        else:
            summary["filtered"] += 1

        if eligible:
            summary["will_be_analyzed"] += 1

    return summary


# ==========================================================
# PÁGINA PRINCIPAL
# ==========================================================

@app.get("/")
def home():

    if AUTH_ENABLED:
        raise HTTPException(
            status_code=404,
            detail="Interface local desabilitada.",
        )

    return FileResponse(
        WEB_DIR / "index.html"
    )


# ==========================================================
# HEALTH CHECK
# ==========================================================
#
# Usado pelo Render (e por qualquer monitor externo) para
# saber se o serviço está de pé. Não consulta banco nem
# JW Player — só confirma que o processo respondeu.

@app.get("/health")
def health():

    return {
        "ok": True
    }


# ==========================================================
# AUTENTICAÇÃO DA APLICAÇÃO
# ==========================================================

@app.post("/api/auth/login")
def app_login(
    credentials: AppLoginRequest,
    request: Request,
):

    if not AUTH_ENABLED:
        return {
            "auth_enabled": False,
            "authenticated": True,
            "username": AUTH_USERNAME or "local",
        }

    if login_is_rate_limited(
        request,
    ):
        raise HTTPException(
            status_code=429,
            detail="Muitas tentativas de login. Tente novamente mais tarde.",
        )

    valid_username = secrets.compare_digest(
        credentials.username.encode("utf-8"),
        AUTH_USERNAME.encode("utf-8"),
    )
    valid_password = secrets.compare_digest(
        credentials.password.encode("utf-8"),
        AUTH_PASSWORD.encode("utf-8"),
    )

    if not valid_username or not valid_password:
        record_login_failure(
            request,
        )
        raise HTTPException(
            status_code=401,
            detail="Usuário ou senha inválidos.",
        )

    clear_login_failures(
        request,
    )

    return {
        "auth_enabled": True,
        "authenticated": True,
        "username": AUTH_USERNAME,
        "access_token": encode_session(AUTH_USERNAME),
        "token_type": "Bearer",
        "expires_in": AUTH_SESSION_TTL_SECONDS,
    }


@app.get("/api/auth/session")
def app_session(
    request: Request,
):

    username = authenticated_username(request)

    return {
        "auth_enabled": AUTH_ENABLED,
        "authenticated": bool(username),
        "username": username or "",
    }


@app.post("/api/auth/logout")
def app_logout():

    global AUTH_SESSION_VERSION

    AUTH_SESSION_VERSION = secrets.token_urlsafe(32)

    return {
        "auth_enabled": AUTH_ENABLED,
        "authenticated": False,
        "username": "",
    }


# ==========================================================
# STATUS DOS SERVIÇOS
# ==========================================================
#
# Nunca retorna chave, token, senha ou cookie — apenas se
# cada integração está configurada/conectada.

@app.get("/api/status")
def service_status():

    session_state = JW_SESSION.status().get(
        "state"
    )

    return {

        "gemini":
            bool(
                AI_API_KEYS.get("Gemini")
            ),

        "claude":
            bool(
                AI_API_KEYS.get("Claude")
            ),

        "ollama_enabled":
            ENABLE_OLLAMA,

        "jw_agent":
            session_state == "connected",
    }


# ==========================================================
# ESTATÍSTICAS
# ==========================================================

@app.get("/api/stats")
def stats():

    library = get_current_library()

    return {

        **DATABASE.stats(),

        "jw_library": library["name"],

        "jw_property_id":
            library["property_id"],

        "jw_library_url":
            library["url"],
    }


# ==========================================================
# VÍDEOS
# ==========================================================

def _dedupe_by_jwplayer_id(
    rows: list[dict],
) -> list[dict]:

    """
    O JWPlayer ID é o identificador único do vídeo. A mesma
    planilha pode trazer o mesmo JWPlayer ID em mais de uma
    linha (IDs de registro diferentes); mantém apenas a
    primeira ocorrência de cada JWPlayer ID.
    """

    seen = set()
    deduped = []

    for row in rows:

        jwplayer_id = row.get("jwplayer_id")

        if jwplayer_id in seen:
            continue

        seen.add(jwplayer_id)
        deduped.append(row)

    return deduped


@app.get("/api/videos")
def videos(
    search: str = "",
    status: str = "",
    category: str = "",
):

    rows = _dedupe_by_jwplayer_id(
        DATABASE.list_portfolio()
    )

    needle = search.casefold().strip()

    if needle:

        rows = [

            row

            for row in rows

            if needle in " ".join(

                str(
                    row.get(key) or ""
                )

                for key in (
                    "lesson_name",
                    "keywords",
                    "jwplayer_id",
                )

            ).casefold()

        ]

    if status:

        rows = [

            row

            for row in rows

            if (
                row.get("status")
                or "Pendente"
            ) == status

        ]

    if category:

        rows = [

            row

            for row in rows

            if row.get(
                "final_category"
            ) == category

        ]

    return {
        "items": rows,
        "total": len(rows),
    }


# ==========================================================
# BACKFILL — PUBLISH DATE DE VÍDEOS JÁ ANALISADOS
# ==========================================================
#
# check_publish_dates() só roda, normalmente, para mídia ainda
# pendente (ver /api/import) — vídeos que já concluíram a
# análise nunca passavam por ali, mesmo que o publish_date
# estivesse ausente. Este endpoint busca somente o metadado de
# publish_date no JW Player para o que estiver faltando na
# execução atual; não chama IA, não refaz transcrição/resumo,
# não altera status/summary/final_category em `analyses`.

@app.post("/api/backfill-publish-dates")
def backfill_publish_dates():

    library = get_current_library()

    rows = _dedupe_by_jwplayer_id(
        DATABASE.list_portfolio()
    )

    missing = [
        row["jwplayer_id"]
        for row in rows
        if not row.get("publish_date")
    ]

    if not missing:
        return {
            "checked": 0,
            "updated": 0,
            "still_missing": 0,
        }

    check_publish_dates(
        missing,
        site_id=library["property_id"],
        min_publish_date="",
    )

    refreshed = {
        row["jwplayer_id"]: row.get("publish_date")
        for row in DATABASE.list_portfolio()
        if row["jwplayer_id"] in set(missing)
    }

    updated = sum(
        1
        for value in refreshed.values()
        if value
    )

    return {
        "checked": len(missing),
        "updated": updated,
        "still_missing": len(missing) - updated,
    }


# ==========================================================
# BACKFILL — CLASSIFICAÇÃO SEMÂNTICA DE VÍDEOS EXISTENTES
# ==========================================================
#
# Classifica (macrotema/microtema/nanotema) vídeos que já
# concluíram a análise e já têm "Resumo do conteúdo" salvo, mas
# ainda não passaram pela classificação semântica — cobre TODO o
# histórico, não só a execução (run_id) atual. Reaproveita o
# resumo já persistido: não baixa o vídeo, não extrai frames,
# não chama a IA multimodal novamente.

@app.post("/api/backfill-classification")
def backfill_classification(
    request: BackfillClassificationRequest,
):

    if request.provider not in AVAILABLE_PROVIDERS:

        raise HTTPException(
            status_code=422,
            detail=AVAILABLE_PROVIDERS_MESSAGE,
        )

    api_key = get_provider_api_key(
        request.provider
    )

    if request.provider == "Ollama":

        try:

            validate_ollama_model(
                request.ollama_url,
                request.model,
            )

        except AIError as exc:

            raise HTTPException(
                status_code=422,
                detail=str(exc),
            ) from exc

    targets = (
        DATABASE.media_pending_topic_classification()
    )

    identified = 0

    for item in targets:

        topics = classify_topics(

            request.provider,

            api_key,

            request.model,

            item["summary"],

            ollama_url=request.ollama_url,
        )

        DATABASE.update_analysis(
            item["jwplayer_id"],
            **topics,
        )

        if topics != DEFAULT_TOPIC_CLASSIFICATION:
            identified += 1

    return {
        "checked": len(targets),
        "identified": identified,
    }


# ==========================================================
# IMPORTAÇÃO DA PLANILHA
# ==========================================================

@app.post("/api/import")
async def import_spreadsheet(
    file: UploadFile = File(...),
):

    content = await file.read()

    try:

        rows, report = read_spreadsheet(
            content,
            file.filename or "planilha.xlsx",
        )

    except SpreadsheetValidationError as exc:

        raise HTTPException(
            status_code=422,
            detail=str(exc),
        ) from exc

    cancel_active_batch()

    outcome = DATABASE.import_rows(
        rows,
        file.filename or "planilha.xlsx",
        # Não usa replace=True (que apaga fisicamente o
        # histórico): cada importação agora gera sua própria
        # execução (run_id) em `imports`, e a Etapa 4 mostra
        # somente a execução mais recente. O histórico
        # permanece no banco para auditoria/consulta.
        replace=False,
    )

    return {
        **report,
        **outcome,
    }


# ==========================================================
# IMPORTAR E PROCESSAR
# ==========================================================

@app.post("/api/import-and-process")
async def import_and_process(
    file: UploadFile = File(...),

    property_id: str = Form(""),

    library: str = Form(DEFAULT_JW_LIBRARY),

    provider: str = Form(
        "Gemini"
    ),

    model: str = Form(
        "gemini-flash-latest"
    ),

    ollama_url: str = Form(
        "http://127.0.0.1:11434"
    ),

    whisper_model: str = Form(
        "small"
    ),

    analysis_mode: str = Form(
        "frames"
    ),

    frame_count: int = Form(
        8
    ),

    min_publish_date: str = Form(
        ""
    ),

    include_missing_date: bool = Form(
        False
    ),
):
    if provider not in AVAILABLE_PROVIDERS:

        raise HTTPException(
            status_code=422,
            detail=AVAILABLE_PROVIDERS_MESSAGE,
        )

    # Validar o formato da data cedo, antes de gastar tempo
    # importando a planilha, se estiver malformada.
    parse_publish_date_cutoff(
        min_publish_date
    )

    library_key = normalize_library(library)
    library_config = JW_LIBRARIES[library_key]

    if (
        property_id
        and property_id.strip()
        != library_config["property_id"]
    ):
        raise HTTPException(
            status_code=422,
            detail=(
                "O Property ID informado não corresponde "
                "à biblioteca selecionada."
            ),
        )

    current_status = JW_SESSION.verify()

    session_property_id = str(
        current_status.get("property_id") or ""
    ).strip()

    if (
        current_status.get("state") == "connected"
        and session_property_id
        and session_property_id != library_config["property_id"]
    ):
        # Mesma conta JW Player, biblioteca diferente: reaproveita
        # a sessão já autenticada só trocando o contexto/property,
        # sem exigir um novo login.
        current_status = JW_SESSION.switch_property(
            library_config["property_id"]
        )

    if current_status.get("state") != "connected":
        raise HTTPException(
            status_code=409,
            detail=(
                "A biblioteca JW Player não está autenticada. "
                "Conecte a biblioteca selecionada antes de enviar a planilha."
            ),
        )

    update_current_library(library_key)

    content = await file.read()

    try:

        rows, report = read_spreadsheet(
            content,
            file.filename or "planilha.xlsx",
        )

    except SpreadsheetValidationError as exc:

        raise HTTPException(
            status_code=422,
            detail=str(exc),
        ) from exc

    cancel_active_batch()

    outcome = DATABASE.import_rows(
        rows,
        file.filename or "planilha.xlsx",
        # Não usa replace=True (que apaga fisicamente o
        # histórico): cada importação agora gera sua própria
        # execução (run_id) em `imports`, e a Etapa 4 mostra
        # somente a execução mais recente. O histórico
        # permanece no banco para auditoria/consulta.
        replace=False,
    )

    imported_media = list(
        dict.fromkeys(
            row["jwplayer_id"]
            for row in rows
        )
    )

    states = {

        item["jwplayer_id"]:
            item["status"]

        for item in DATABASE.unique_media()

    }

    pending_media = [

        media_id

        for media_id in imported_media

        if states.get(
            media_id
        ) != "Concluído"

    ]

    # --------------------------------------------------------
    # FILTRO DE PUBLICAÇÃO — ANTES de qualquer processamento
    # --------------------------------------------------------
    #
    # A planilha NÃO inicia mais a análise automaticamente.
    # Esta etapa apenas consulta o Publish date de cada vídeo
    # pendente e marca elegibilidade — nenhum vídeo é baixado,
    # nenhum frame é extraído, nenhuma IA é chamada aqui. O
    # usuário revisa o resumo e confirma em /api/start-eligible.

    filter_summary = check_publish_dates(
        pending_media,
        site_id=library_config["property_id"],
        min_publish_date=min_publish_date,
        include_missing_date=include_missing_date,
    )

    return {

        **report,

        **outcome,

        "pending_media":
            len(pending_media),

        "filter":
            filter_summary,

        "library":
            library_key,

        "property_id":
            library_config["property_id"],
    }


# ==========================================================
# INICIAR ANÁLISE (SOMENTE VÍDEOS ELEGÍVEIS)
# ==========================================================
#
# Botão "Iniciar análise": só é chamado depois que o usuário
# revisou o resumo do filtro de publicação. Enfileira apenas
# os JWPlayer IDs já marcados eligible_for_analysis=1 na
# execução (run_id) confirmada — vídeos filtrados/sem data/erro
# nunca chegam a este ponto.

@app.post("/api/start-eligible")
def start_eligible(
    request: StartEligibleRequest,
):

    states = {

        item["jwplayer_id"]:
            item["status"]

        for item in DATABASE.unique_media()

    }

    eligible_media = [

        jwplayer_id

        for jwplayer_id in DATABASE.eligible_media_for_run(
            request.run_id
        )

        if states.get(
            jwplayer_id
        ) != "Concluído"

    ]

    if not eligible_media:

        return {
            "jobs": [],
            "media_count": 0,
            "message": (
                "Nenhum vídeo elegível pendente "
                "para iniciar análise."
            ),
        }

    process_request = build_process_request(

        media_ids=eligible_media,

        provider=request.provider,

        model=request.model,

        ollama_url=request.ollama_url,

        whisper_model=request.whisper_model,

        analysis_mode=request.analysis_mode,

        frame_count=request.frame_count,
    )

    queued = enqueue_jobs(
        process_request
    )

    return {
        "jobs": queued,
        "media_count": len(eligible_media),
    }


# ==========================================================
# LOGIN JW PLAYER
# ==========================================================

@app.post("/api/jw/login")
def jw_login(
    request: LoginRequest,
):
    global CURRENT_JW_LIBRARY
    global CURRENT_JW_PROPERTY_ID

    library_key = normalize_library(request.library)
    library = JW_LIBRARIES[library_key]

    sent_property_id = str(
        request.property_id or ""
    ).strip()

    if (
        sent_property_id
        and sent_property_id != library["property_id"]
    ):
        raise HTTPException(
            status_code=422,
            detail=(
                "O Property ID informado não corresponde "
                "à biblioteca selecionada."
            ),
        )

    try:
        update_current_library(library_key)

        result = JW_SESSION.login(
            email=request.email,
            password=request.password,
            property_id=library["property_id"],
        )

        result = dict(result or {})
        result["library"] = library_key
        result["library_name"] = library["name"]
        result["property_id"] = library["property_id"]
        result["library_url"] = library["url"]
        result["connected"] = (
            result.get("state") == "connected"
            or result.get("connected") is True
        )

        return result

    except JWSessionError as exc:
        raise HTTPException(
            status_code=409,
            detail=str(exc),
        ) from exc

    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Não foi possível conectar ao JW Player: {exc}",
        ) from exc


# ==========================================================
# TROCAR DE BIBLIOTECA JW PLAYER (MESMA SESSÃO)
# ==========================================================
#
# As bibliotecas pertencem à mesma conta JW Player — trocar de
# biblioteca é apenas mudar o contexto/property da pesquisa, não
# uma nova autenticação. Reaproveita a sessão Playwright já
# conectada (JW_SESSION.switch_property): não recebe nem exige
# e-mail/senha. Só resulta em desconectado/atenção quando a
# sessão realmente não está mais válida para nenhuma biblioteca.

@app.post("/api/jw/switch-library")
def jw_switch_library(
    request: SwitchLibraryRequest,
):
    library_key = normalize_library(request.library)
    library = JW_LIBRARIES[library_key]

    sent_property_id = str(
        request.property_id or ""
    ).strip()

    if (
        sent_property_id
        and sent_property_id != library["property_id"]
    ):
        raise HTTPException(
            status_code=422,
            detail=(
                "O Property ID informado não corresponde "
                "à biblioteca selecionada."
            ),
        )

    update_current_library(library_key)

    try:
        result = JW_SESSION.switch_property(
            library["property_id"]
        )

    except JWSessionError as exc:
        raise HTTPException(
            status_code=409,
            detail=str(exc),
        ) from exc

    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail=(
                "Não foi possível trocar de biblioteca "
                f"no JW Player: {exc}"
            ),
        ) from exc

    result = dict(result or {})
    result["library"] = library_key
    result["library_name"] = library["name"]
    result["property_id"] = library["property_id"]
    result["library_url"] = library["url"]
    result["connected"] = (
        result.get("state") == "connected"
        or result.get("connected") is True
    )

    return result


# ==========================================================
# STATUS JW PLAYER
# ==========================================================

@app.get("/api/jw/status")
def jw_status(
    property_id: str = "",
    library: str = "",
    verify: bool = True,
):
    requested_library_key = (
        normalize_library(library)
        if library
        else CURRENT_JW_LIBRARY
    )

    requested_library = JW_LIBRARIES[
        requested_library_key
    ]

    if (
        property_id
        and property_id.strip()
        != requested_library["property_id"]
    ):
        raise HTTPException(
            status_code=422,
            detail=(
                "O Property ID não corresponde à "
                "biblioteca selecionada."
            ),
        )

    try:
        status = (
            JW_SESSION.verify()
            if verify
            else JW_SESSION.status()
        )
    except JWSessionError as exc:
        status = {
            "state": "error",
            "connected": False,
            "message": str(exc),
            "property_id": "",
        }

    status = dict(status or {})

    session_property_id = str(
        status.get("property_id") or ""
    ).strip()

    session_library_key = library_from_property_id(
        session_property_id
    )

    status["selected_library"] = requested_library_key
    status["selected_property_id"] = (
        requested_library["property_id"]
    )
    status["library"] = requested_library_key
    status["library_name"] = requested_library["name"]
    status["library_url"] = requested_library["url"]

    if (
        status.get("state") == "connected"
        and session_property_id
        and session_property_id
        != requested_library["property_id"]
    ):
        status["state"] = "disconnected"
        status["connected"] = False
        status["message"] = (
            "A sessão JW Player está conectada à "
            f"biblioteca {session_library_key or session_property_id}, "
            f"mas a biblioteca selecionada é "
            f"{requested_library['name']}."
        )

    status["property_id"] = (
        session_property_id
        or requested_library["property_id"]
    )

    return status


# ==========================================================
# ATUALIZAR JOB
# ==========================================================

def update_job(
    job_id: str,
    **values,
) -> None:

    with JOBS_LOCK:

        if job_id in JOBS:

            JOBS[job_id].update(
                values
            )


def _prune_jobs_locked() -> None:

    """
    Remove as entradas mais antigas de JOBS quando ultrapassa
    MAX_JOBS_HISTORY. O chamador precisa já estar de posse de
    JOBS_LOCK — dicts em Python preservam ordem de inserção,
    então isso remove sempre os jobs criados há mais tempo.
    """

    while len(JOBS) > MAX_JOBS_HISTORY:

        oldest_job_id = next(
            iter(JOBS)
        )

        JOBS.pop(
            oldest_job_id,
            None,
        )


# ==========================================================
# CANCELAR LOTE ATIVO
# ==========================================================

def cancel_active_batch() -> None:

    global ACTIVE_BATCH

    with ACTIVE_BATCH_LOCK:

        if ACTIVE_BATCH is not None:

            ACTIVE_BATCH.cancelled = True

            ACTIVE_BATCH.reason = (
                "Lote substituído por uma nova planilha."
            )

            ACTIVE_BATCH.set()

        ACTIVE_BATCH = None

    with JOBS_LOCK:

        JOBS.clear()


# ==========================================================
# VALIDAR LOTE
# ==========================================================

def ensure_batch_active(
    batch_abort: threading.Event,
) -> None:

    if getattr(
        batch_abort,
        "cancelled",
        False,
    ):

        raise BatchCancelled(
            "Lote substituído por uma nova planilha."
        )


# ==========================================================
# INTERRUPÇÃO JW PLAYER
# ==========================================================

def is_session_interruption(
    exc: Exception,
) -> bool:

    if not isinstance(
        exc,
        JWSessionError,
    ):

        return False

    message = str(
        exc
    ).casefold()

    return any(

        marker in message

        for marker in (

            "conecte uma sessão",

            "navegador jw player foi fechado",

            "sessão jw player expirou",

            "sessão expirou",

            "autenticação",

        )

    )


# ==========================================================
# EXECUTAR JOB
# ==========================================================

def run_media_job(
    job_id: str,
    media_id: str,
    request: ProcessRequest,
    batch_abort: threading.Event,
) -> None:

    with ANALYSIS_LOCK:

        _run_media_job_serial(

            job_id,

            media_id,

            request,

            batch_abort,
        )


# ==========================================================
# EXECUÇÃO SERIAL
# ==========================================================

def _run_media_job_serial(
    job_id: str,
    media_id: str,
    request: ProcessRequest,
    batch_abort: threading.Event,
) -> None:

    if getattr(
        batch_abort,
        "cancelled",
        False,
    ):

        update_job(

            job_id,

            state="cancelled",

            stage="Cancelado",

            message=(
                "Lote substituído "
                "por uma nova planilha."
            ),
        )

        return

    if batch_abort.is_set():

        reason = getattr(

            batch_abort,

            "reason",

            (
                "Corrija a conexão "
                "ou configuração "
                "e tente novamente."
            ),
        )

        DATABASE.update_analysis(

            media_id,

            status="Pendente",

            error_message=None,
        )

        update_job(

            job_id,

            state="paused",

            stage="Aguardando",

            message=(
                f"Lote pausado: {reason}"
            ),
        )

        return

    update_job(

        job_id,

        state="running",

        stage="Validando vídeo",

        message=(
            "Capturando o vídeo no JW Player"
        ),
    )

    DATABASE.update_analysis(

        media_id,

        status="Processando",

        error_message=None,
    )

    work_dir = (
        DATA_DIR
        / "work"
        / media_id
    )

    try:

        # --------------------------------------------------
        # CAPTURAR VÍDEO
        # --------------------------------------------------

        captured = JW_SESSION.capture_media(
            media_id
        )

        ensure_batch_active(
            batch_abort
        )

        master_url = (
            captured.get(
                "master_url"
            )
            or captured.get(
                "url"
            )
        )

        if not master_url:

            raise JWSessionError(
                "Não foi possível obter a URL do vídeo no JW Player."
            )

        update_job(

            job_id,

            stage="Extraindo informações",

            message=(
                "Vídeo validado. "
                "Obtendo informações do JW Player"
            ),
        )

        # --------------------------------------------------
        # FRAMES
        # --------------------------------------------------

        update_job(

            job_id,

            stage="Extraindo frames",

            message=(
                f"Extraindo "
                f"{request.frame_count} "
                f"frames distribuídos"
            ),
        )

        frames, duration = extract_frames(

            master_url,

            work_dir,

            request.frame_count,
        )

        ensure_batch_active(
            batch_abort
        )

        # --------------------------------------------------
        # TRANSCRIÇÃO
        # --------------------------------------------------

        transcript = ""

        if request.analysis_mode == "hybrid":

            update_job(

                job_id,

                stage="Transcrevendo áudio",

                message=(
                    "Transcrevendo áudio "
                    "complementar "
                    "(modo híbrido)"
                ),
            )

            transcript = transcribe_hls(

                master_url,

                work_dir,

                request.whisper_model,
            )

            ensure_batch_active(
                batch_abort
            )

        # --------------------------------------------------
        # TÍTULO
        # --------------------------------------------------

        title = next(

            (

                item["lesson_name"]

                for item
                in DATABASE.unique_media()

                if item["jwplayer_id"]
                == media_id

            ),

            media_id,
        )

        # --------------------------------------------------
        # IA
        # --------------------------------------------------

        update_job(

            job_id,

            stage="Analisando conteúdo visual",

            message=(
                f"Classificando e "
                f"resumindo com "
                f"{request.provider}"
            ),
        )

        if (
            request.provider != "Ollama"
            and not request.api_key
        ):

            raise AIError(

                f"A chave da "
                f"{request.provider} "
                f"não está configurada "
                f"no servidor."
            )

        result = analyze_frames(

            request.provider,

            request.api_key,

            request.model,

            title,

            frames,

            transcript=transcript,

            ollama_url=request.ollama_url,
        )

        ensure_batch_active(
            batch_abort
        )

        update_job(

            job_id,

            stage="Confirmando professor",

            message=(
                "Cruzando evidências visuais "
                "e de fala para confirmar o professor"
            ),
        )

        update_job(

            job_id,

            stage="Gerando resumo",

            message=(
                "Gerando resumo e nome da aula"
            ),
        )

        # --------------------------------------------------
        # CLASSIFICAÇÃO SEMÂNTICA (MACRO/MICRO/NANOTEMA)
        # --------------------------------------------------
        #
        # Etapa separada, a partir do resumo já gerado acima —
        # não reenvia frames nem reprocessa o vídeo. Nunca
        # levanta exceção (classify_topics já trata falhas
        # internamente), então não pode pausar o lote nem
        # marcar este vídeo como erro.

        update_job(

            job_id,

            stage="Classificando tema",

            message=(
                "Identificando macrotema, "
                "microtema e nanotema"
            ),
        )

        topics = classify_topics(

            request.provider,

            request.api_key,

            request.model,

            result["summary"],

            ollama_url=request.ollama_url,
        )

        # --------------------------------------------------
        # SALVAR RESULTADO
        # --------------------------------------------------

        DATABASE.update_analysis(

            media_id,

            status="Concluído",

            ai_category=result[
                "category"
            ],

            final_category=result[
                "category"
            ],

            summary=result[
                "summary"
            ],

            confidence=result[
                "confidence"
            ],

            professor_name=result[
                "professor_name"
            ],

            validation_status="Pendente",

            transcript=transcript,

            source_title=title,

            duration=duration,

            analyzed_at=utc_now(),

            error_message=None,

            macrotema=topics[
                "macrotema"
            ],

            microtema=topics[
                "microtema"
            ],

            nanotema=topics[
                "nanotema"
            ],
        )

        update_job(

            job_id,

            state="completed",

            stage="Concluído",

            message=(
                "Processamento concluído"
            ),

            result={

                **result,

                **topics,

                "title":
                    title,

                "duration":
                    duration,
            },
        )

    except BatchCancelled:

        update_job(

            job_id,

            state="cancelled",

            stage="Cancelado",

            message=(
                "Lote substituído "
                "por uma nova planilha."
            ),
        )

    except Exception as exc:

        message = str(
            exc
        )

        if is_session_interruption(
            exc
        ):

            batch_abort.reason = (

                "A sessão do JW Player "
                "foi interrompida. "
                "Reconecte para retomar."
            )

            batch_abort.set()

            DATABASE.update_analysis(

                media_id,

                status="Pendente",

                error_message=message,
            )

            update_job(

                job_id,

                state="paused",

                stage="Aguardando",

                message=(

                    "Processamento pausado. "
                    "Reconecte o JW Player "
                    "e envie novamente "
                    "a planilha para retomar."
                ),
            )

        elif isinstance(
            exc,
            AIResponseError,
        ):

            # A IA respondeu, mas o texto não pôde ser
            # interpretado como o JSON esperado (já foi
            # tentado novamente algumas vezes dentro de
            # analyze_frames, sem sucesso). Isso é uma falha
            # deste vídeo específico, não da configuração do
            # provedor — por isso NÃO pausa o lote: apenas
            # este vídeo fica com status "Erro" e o próximo
            # vídeo da fila continua normalmente.

            logger.warning(
                "Resposta da IA nao pode ser interpretada "
                "como JSON | job_id=%s jwplayer_id=%s "
                "provider=%s model=%s erro=%s "
                "resposta_bruta=%r",
                job_id,
                media_id,
                request.provider,
                request.model,
                message,
                exc.raw_text,
            )

            DATABASE.update_analysis(

                media_id,

                status="Erro",

                error_message=message,
            )

            update_job(

                job_id,

                state="error",

                stage="Erro",

                message=(
                    "Resposta da IA inválida."
                ),
            )

        elif isinstance(
            exc,
            AIError,
        ):

            batch_abort.reason = message

            batch_abort.set()

            DATABASE.update_analysis(

                media_id,

                status="Pendente",

                error_message=message,
            )

            update_job(

                job_id,

                state="paused",

                stage="Aguardando",

                message=(

                    "Lote pausado por "
                    "configuração da IA: "
                    f"{message}"
                ),
            )

        else:

            DATABASE.update_analysis(

                media_id,

                status="Erro",

                error_message=message,
            )

            update_job(

                job_id,

                state="error",

                stage="Erro",

                message=message,
            )

    finally:

        request.api_key = ""

        # Frames (base64), transcrição e o modelo Whisper usado no
        # modo híbrido geram picos grandes de memória por vídeo.
        # Forçar a coleta aqui, ao final de cada job (sucesso ou
        # falha), evita que esses picos se acumulem entre um vídeo
        # e o próximo dentro do mesmo processo de vida longa.
        gc.collect()


# ==========================================================
# ENFILEIRAR JOBS
# ==========================================================

def enqueue_jobs(
    request: ProcessRequest,
) -> list[dict]:

    global ACTIVE_BATCH

    if request.provider not in AVAILABLE_PROVIDERS:

        raise HTTPException(

            status_code=422,

            detail=AVAILABLE_PROVIDERS_MESSAGE,
        )

    current_status = JW_SESSION.status()

    if current_status.get(
        "state"
    ) != "connected":

        raise HTTPException(

            status_code=409,

            detail=(
                "Conecte o JW Player "
                "antes de processar."
            ),
        )

    request.api_key = (
        get_provider_api_key(
            request.provider
        )
    )

    if request.provider == "Ollama":

        try:

            validate_ollama_model(

                request.ollama_url,

                request.model,
            )

        except AIError as exc:

            raise HTTPException(

                status_code=422,

                detail=str(exc),

            ) from exc

    jobs = []

    batch_abort = (
        threading.Event()
    )

    batch_abort.cancelled = False

    with ACTIVE_BATCH_LOCK:

        ACTIVE_BATCH = (
            batch_abort
        )

    # NÃO chamar DATABASE.ensure_video_for_jwplayer_id() aqui
    # dentro do loop: essa função cria uma NOVA execução
    # (run_id) a cada chamada. Se fosse chamada uma vez por
    # media_id, uma planilha com várias linhas fragmentaria a
    # própria execução em uma linha por vídeo, e a Etapa 4
    # (filtrada pela execução mais recente) passaria a mostrar
    # somente o último vídeo da lista, perdendo os demais
    # resultados já concluídos. Cada chamador de enqueue_jobs()
    # é responsável por garantir a linha/execução UMA vez para
    # todo o conjunto que está enviando (import_rows() já faz
    # isso para a planilha; analyze_jwplayer() faz isso para a
    # análise individual).

    for media_id in dict.fromkeys(
        request.media_ids
    ):

        job_id = uuid.uuid4().hex

        job_data = {

            "id": job_id,

            "media_id": media_id,

            "jwplayer_id": media_id,

            "state": "queued",

            "stage": "Aguardando",

            "message": "Na fila",

            "provider": request.provider,

            "model": request.model,

            "created_at": utc_now(),

        }

        with JOBS_LOCK:

            JOBS[job_id] = (
                job_data
            )

            _prune_jobs_locked()

        PROCESSOR.submit(

            run_media_job,

            job_id,

            media_id,

            request.model_copy(
                deep=True
            ),

            batch_abort,
        )

        jobs.append(
            job_data
        )

    request.api_key = ""

    return jobs


# ==========================================================
# ANALISAR JWPLAYER ID INDIVIDUAL
# ==========================================================

@app.post("/api/analyze-jwplayer")
def analyze_jwplayer(
    request: AnalyzeJWPlayerRequest,
):

    jwplayer_id = str(
        request.jwplayer_id or ""
    ).strip()

    if not jwplayer_id:

        raise HTTPException(

            status_code=422,

            detail=(
                "Informe o JWPlayer ID "
                "do vídeo."
            ),
        )

    library_key = normalize_library(
        request.library
    )

    library = JW_LIBRARIES[
        library_key
    ]

    # ------------------------------------------------------
    # VALIDAR PROPERTY ID
    # ------------------------------------------------------

    if request.property_id:

        sent_property_id = str(
            request.property_id
        ).strip()

        if (
            sent_property_id
            != library["property_id"]
        ):

            raise HTTPException(

                status_code=422,

                detail=(
                    "O Property ID informado "
                    "não corresponde à "
                    "biblioteca selecionada."
                ),
            )

    # ------------------------------------------------------
    # VALIDAR SESSÃO
    # ------------------------------------------------------

    session_status = (
        JW_SESSION.status()
    )

    session_property_id = str(

        session_status.get(
            "property_id"
        )

        or session_status.get(
            "propertyId"
        )

        or ""

    ).strip()

    # Mesma conta JW Player, biblioteca diferente: reaproveita a
    # sessão já autenticada só trocando o contexto/property, sem
    # exigir um novo login.
    if (
        session_status.get("state") == "connected"
        and session_property_id
        and session_property_id != library["property_id"]
    ):

        session_status = JW_SESSION.switch_property(
            library["property_id"]
        )

    if session_status.get(
        "state"
    ) != "connected":

        raise HTTPException(

            status_code=409,

            detail=(
                "Conecte primeiro "
                "a biblioteca JW Player."
            ),
        )

    # ------------------------------------------------------
    # GARANTIR BIBLIOTECA ATUAL
    # ------------------------------------------------------

    update_current_library(
        library_key
    )

    # ------------------------------------------------------
    # CRIAR REQUEST
    # ------------------------------------------------------

    process_request = (
        build_process_request(

            media_ids=[
                jwplayer_id
            ],

            provider=request.provider,

            model=request.model,

            ollama_url=request.ollama_url,

            whisper_model=request.whisper_model,

            analysis_mode=request.analysis_mode,

            frame_count=request.frame_count,

        )
    )

    # ------------------------------------------------------
    # GARANTIR VÍDEO/EXECUÇÃO
    # ------------------------------------------------------

    # Uma análise individual é a sua própria execução (run_id).
    # Chamado uma única vez aqui — não dentro de enqueue_jobs(),
    # que é compartilhado com a planilha (que já garante sua
    # própria execução via DATABASE.import_rows()).
    DATABASE.ensure_video_for_jwplayer_id(
        jwplayer_id
    )

    # ------------------------------------------------------
    # FILTRO DE PUBLICAÇÃO
    # ------------------------------------------------------
    #
    # A análise individual respeita o mesmo filtro da planilha
    # — não é uma exceção silenciosa. Sem min_publish_date
    # informado, todo vídeo é elegível (comportamento anterior
    # preservado).

    filter_summary = check_publish_dates(
        [jwplayer_id],
        site_id=library["property_id"],
        min_publish_date=request.min_publish_date,
        include_missing_date=request.include_missing_date,
    )

    if filter_summary["will_be_analyzed"] == 0:

        with DATABASE.connect() as connection:

            filtered_row = connection.execute(
                """
                SELECT publish_date, filter_status, filter_reason

                FROM videos

                WHERE jwplayer_id = ?

                LIMIT 1
                """,
                (jwplayer_id,),
            ).fetchone()

        raise HTTPException(

            status_code=422,

            detail=(

                "Este vídeo não é elegível para análise "
                "pelo filtro de data de publicação "
                f"(Publish date: "
                f"{filtered_row['publish_date'] if filtered_row else 'não encontrado'}, "
                f"motivo: "
                f"{filtered_row['filter_reason'] if filtered_row else 'desconhecido'})."
            ),
        )

    # ------------------------------------------------------
    # ENFILEIRAR
    # ------------------------------------------------------

    queued = enqueue_jobs(
        process_request
    )

    return {

        "ok": True,

        "message": (
            f"JWPlayer ID "
            f"{jwplayer_id} "
            "enviado para análise."
        ),

        "jwplayer_id":
            jwplayer_id,

        "library":
            library_key,

        "library_name":
            library["name"],

        "property_id":
            library["property_id"],

        "jobs":
            queued,
    }


# ==========================================================
# PROCESSAMENTO DIRETO
# ==========================================================

@app.post("/api/process")
def process(
    request: ProcessRequest,
):

    request.api_key = (
        get_provider_api_key(
            request.provider
        )
    )

    return {
        "jobs":
            enqueue_jobs(
                request
            )
    }


# ==========================================================
# JOBS
# ==========================================================

@app.get("/api/jobs")
def jobs():

    with JOBS_LOCK:

        return {
            "items":
                list(
                    JOBS.values()
                )[-100:]
        }


# ==========================================================
# VALIDAÇÃO
# ==========================================================

@app.post("/api/validate")
def validate(
    request: ValidationRequest,
):

    DATABASE.update_analysis(

        request.jwplayer_id,

        final_category=(
            request.final_category
        ),

        summary=(
            " ".join(
                request.summary.split()
            )
        ),

        validation_status=(
            "Validado"
            if request.validated
            else "Pendente"
        ),
    )

    return {
        "ok": True
    }


# ==========================================================
# EXPORTAÇÃO CSV
# ==========================================================

def _publish_date_year(value) -> int | None:

    """
    O publish_date é armazenado como ISO 8601 (podendo conter
    horário e timezone, ex.: 2025-05-13T11:00:18+00:00). Extrai
    apenas o ano, sem alterar/descartar o valor original.
    """

    text = str(
        value or ""
    ).strip()

    if not text:
        return None

    normalized = (
        text[:-1] + "+00:00"
        if text.endswith("Z")
        else text
    )

    try:
        return datetime.fromisoformat(
            normalized
        ).year
    except ValueError:
        pass

    try:
        return date.fromisoformat(
            text[:10]
        ).year
    except ValueError:
        return None


# Fonte única do "ano de publicação": usada pelo filtro da tela
# (via /api/videos + publishDateYear() em app.js), pela
# exportação CSV, pela exportação XLSX e pelo resumo por ano —
# nenhuma dessas quatro pontas reimplementa a extração do ano.
#
# Ordem e nomes das colunas seguem a especificação de
# exportação (planilha de referência). Não existe coluna própria
# para o link da aula: no XLSX, a própria célula "JWPlayer ID"
# recebe o hyperlink (ver export_xlsx) — mesma regra do site,
# via build_jwplayer_media_url().
EXPORT_COLUMNS = (
    ("lesson_name", "Nome da aula"),
    ("final_category", "Modelo de aula"),
    ("professor_name", "Professor"),
    ("summary", "Resumo do conteúdo"),
    ("jwplayer_id", "JWPlayer ID"),
    ("status", "Status"),
    ("validation_status", "Validação Manual"),
    ("confidence", "Confiança"),
    ("keywords", "Palavras-chave"),
    ("publish_year", "Ano de publicação"),
    ("macrotema", "Macrotema"),
    ("microtema", "Microtema"),
    ("nanotema", "Nanotema"),
)


def _rows_for_export(
    year: str = "",
) -> list[dict]:

    """
    Fonte única de linhas para CSV e XLSX:
    - dedupe por JWPlayer ID (identificador único do vídeo);
    - somente status "Concluído";
    - ano de publicação calculado uma única vez por linha, com
      a mesma lógica usada no filtro da tela
      (_publish_date_year), reaproveitado tanto para o valor
      exportado quanto para o resumo por ano.
    """

    rows = _dedupe_by_jwplayer_id(
        DATABASE.list_portfolio()
    )

    # Regra obrigatória: a exportação só pode conter aulas com
    # status "Concluído" — o filtro é aplicado antes de gerar
    # o arquivo, independente de outros filtros da tela.
    rows = [
        dict(row)
        for row in rows
        if row.get("status") == "Concluído"
    ]

    library = get_current_library()

    for row in rows:

        export_year = _publish_date_year(
            row.get("publish_date")
        )

        row["_export_year"] = export_year

        row["publish_year"] = (
            str(export_year)
            if export_year is not None
            else ""
        )

        # Mesma regra de link usada pelo site
        # (buildJWPlayerMediaUrl em web/app.js), guardada aqui
        # como campo interno (não é uma coluna de EXPORT_COLUMNS):
        # usada só pelo XLSX para aplicar o hyperlink diretamente
        # na célula "JWPlayer ID" — ver export_xlsx(). Se não for
        # possível gerar a URL, fica vazio sem interromper a
        # exportação dos demais registros.
        try:

            row["_jwplayer_link"] = build_jwplayer_media_url(
                row.get("jwplayer_id"),
                library["property_id"],
            )

        except Exception:

            logger.warning(
                "Falha ao gerar o link do JWPlayer ID na "
                "exportação | jwplayer_id=%s",
                row.get("jwplayer_id"),
            )

            row["_jwplayer_link"] = ""

        for field in (
            "macrotema",
            "microtema",
            "nanotema",
        ):

            row[field] = (
                row.get(field)
                or "Não identificado"
            )

    year = str(
        year or ""
    ).strip()

    if year:

        try:
            year_number = int(year)
        except ValueError:
            raise HTTPException(
                status_code=422,
                detail=(
                    "Ano de publicação inválido."
                ),
            )

        rows = [
            row
            for row in rows
            if row["_export_year"] == year_number
        ]

    return rows


@app.get("/api/export.csv")
def export_csv(year: str = ""):

    output = io.StringIO()

    fields = [key for key, _ in EXPORT_COLUMNS]

    writer = csv.DictWriter(

        output,

        fieldnames=fields,

        extrasaction="ignore",
    )

    writer.writerow(
        dict(EXPORT_COLUMNS)
    )

    writer.writerows(
        _rows_for_export(year=year)
    )

    return Response(

        content=(
            "\ufeff"
            + output.getvalue()
        ),

        media_type=(
            "text/csv; charset=utf-8"
        ),

        headers={

            "Content-Disposition":
                (
                    'attachment; '
                    'filename='
                    '"portfolio_cetrus.csv"'
                )
        },
    )


@app.get("/api/export.xlsx")
def export_xlsx(year: str = ""):

    rows = _rows_for_export(year=year)

    workbook = Workbook()

    videos_sheet = workbook.active
    videos_sheet.title = "Vídeos"

    videos_sheet.append(
        [label for _, label in EXPORT_COLUMNS]
    )

    jwplayer_id_column = (
        [key for key, _ in EXPORT_COLUMNS].index(
            "jwplayer_id"
        )
        + 1
    )

    for row in rows:

        videos_sheet.append([

            row["_export_year"]
            if key == "publish_year"
            else row.get(key, "")

            for key, _ in EXPORT_COLUMNS

        ])

        # A própria célula "JWPlayer ID" recebe o hyperlink —
        # não existe coluna separada de link. O texto exibido
        # continua sendo somente o JWPlayer ID (já preenchido
        # pelo append acima); aponta para a mesma URL usada pelo
        # site — ver build_jwplayer_media_url().
        link_url = row.get(
            "_jwplayer_link"
        )

        if link_url:

            link_cell = videos_sheet.cell(
                row=videos_sheet.max_row,
                column=jwplayer_id_column,
            )

            link_cell.hyperlink = link_url

            link_cell.font = Font(
                color="0563C1",
                underline="single",
            )

    last_column = get_column_letter(
        len(EXPORT_COLUMNS)
    )

    videos_sheet.auto_filter.ref = (
        f"A1:{last_column}{len(rows) + 1}"
    )

    summary_sheet = workbook.create_sheet(
        "Resumo por ano"
    )

    summary_sheet.append(
        ["Ano de publicação", "Total de vídeos"]
    )

    counts: dict[str, int] = {}

    for row in rows:

        export_year = row["_export_year"]

        label = (
            str(export_year)
            if export_year is not None
            else "Não informado"
        )

        counts[label] = (
            counts.get(label, 0) + 1
        )

    def _summary_sort_key(label: str):
        return (
            (0, int(label))
            if label.isdigit()
            else (1, label)
        )

    for label in sorted(
        counts,
        key=_summary_sort_key,
    ):
        summary_sheet.append(
            [label, counts[label]]
        )

    buffer = io.BytesIO()
    workbook.save(buffer)

    return Response(

        content=buffer.getvalue(),

        media_type=(
            "application/vnd.openxmlformats-officedocument"
            ".spreadsheetml.sheet"
        ),

        headers={

            "Content-Disposition":
                (
                    'attachment; '
                    'filename='
                    '"portfolio_cetrus.xlsx"'
                )
        },
    )
