import csv
import io
import unittest
from pathlib import Path
from unittest.mock import patch

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app import (
    AI_API_KEYS,
    ANALYSIS_LOCK,
    DATABASE,
    EXPORT_COLUMNS,
    JW_SESSION,
    PROCESSOR,
    _dedupe_by_jwplayer_id,
    _publish_date_year,
    _rows_for_export,
    app,
    build_jwplayer_media_url,
    get_current_library,
    is_session_interruption,
)
from src.portfolio.ai import AIError
from src.portfolio.database import Database
from src.portfolio.jw_session import JWSessionError

FORBIDDEN_CSV_COLUMNS = {
    "Motivo do filtro",
    "Elegível para análise",
    "Filtro de data",
    "Publish date",
    "Duração",
    "Link da aula",
}


class ApiTests(unittest.TestCase):
    def setUp(self):
        self.client = TestClient(app)

    def test_home_and_portfolio_endpoints(self):
        home = self.client.get("/")
        self.assertEqual(home.status_code, 200)
        self.assertIn("Portfólio de vídeos Cetrus", home.text)
        stats = self.client.get("/api/stats")
        self.assertEqual(stats.status_code, 200)
        self.assertIn("records", stats.json())
        self.assertEqual(
            stats.json()["jw_library_url"],
            "https://dashboard.jwplayer.com/p/XdfUPSCL/media",
        )
        videos = self.client.get("/api/videos", params={"search": "y6zIncnf"})
        self.assertEqual(videos.status_code, 200)
        self.assertIn("items", videos.json())

    def test_video_processor_is_strictly_sequential(self):
        self.assertEqual(PROCESSOR._max_workers, 1)
        self.assertFalse(ANALYSIS_LOCK.locked())

    def test_processing_requires_authenticated_jw_session(self):
        # Claude precisa de uma chave configurada (AI_API_KEYS) para
        # passar da validação de provedor/chave e chegar à validação
        # de sessão JW Player, que é o que este teste verifica.
        # Sem isso, a requisição falha antes com 422 (chave ausente),
        # independente da sessão estar conectada ou não.
        with patch.dict(AI_API_KEYS, {"Claude": "test-key"}):
            response = self.client.post(
                "/api/process",
                json={"media_ids": ["y6zIncnf"], "provider": "Claude", "api_key": ""},
            )
        self.assertEqual(response.status_code, 409)
        self.assertEqual(
            response.json()["detail"],
            "Conecte o JW Player antes de processar.",
        )

    def test_processing_rejects_removed_ai_providers(self):
        response = self.client.post(
            "/api/process",
            json={"media_ids": ["y6zIncnf"], "provider": "OpenAI", "api_key": "key"},
        )
        self.assertEqual(response.status_code, 422)
        self.assertEqual(
            response.json()["detail"],
            "Selecione Gemini, Claude ou Ollama.",
        )

    # TESTE 1/2/3 (troca de biblioteca) — reaproveita a sessão já
    # conectada, sem pedir e-mail/senha novamente.
    def test_switch_library_reuses_session_without_credentials(self):
        with patch.object(
            JW_SESSION,
            "switch_property",
            return_value={
                "state": "connected",
                "connected": True,
                "property_id": "FvJr6FNj",
                "message": "Sessão JW Player conectada e autenticada.",
            },
        ) as mock_switch:
            response = self.client.post(
                "/api/jw/switch-library",
                json={"library": "DEV", "property_id": "FvJr6FNj"},
            )

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body["connected"])
        self.assertEqual(body["library"], "DEV")
        self.assertEqual(body["property_id"], "FvJr6FNj")

        # Nenhuma credencial é enviada nem exigida para trocar de
        # biblioteca — só o property_id da nova biblioteca.
        mock_switch.assert_called_once_with("FvJr6FNj")

    def test_switch_library_rejects_mismatched_property_id(self):
        response = self.client.post(
            "/api/jw/switch-library",
            json={"library": "DEV", "property_id": "XdfUPSCL"},
        )
        self.assertEqual(response.status_code, 422)

    # TESTE 4 — sessão realmente expirada: a troca reporta
    # desconectado (permitindo um novo login normal pela tela),
    # em vez de levantar um erro.
    def test_switch_library_reports_expired_session_without_error(self):
        with patch.object(
            JW_SESSION,
            "switch_property",
            return_value={
                "state": "disconnected",
                "connected": False,
                "property_id": "",
                "message": (
                    "A sessão JW Player expirou. Conecte novamente "
                    "para acessar essa biblioteca."
                ),
            },
        ):
            response = self.client.post(
                "/api/jw/switch-library",
                json={"library": "EBSERH", "property_id": "UBP82vRQ"},
            )

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertFalse(body["connected"])
        self.assertEqual(body["state"], "disconnected")

    # A análise individual não deve exigir reconexão só porque a
    # sessão está conectada a uma biblioteca diferente da
    # selecionada — deve reaproveitar a sessão (switch_property).
    def test_analyze_jwplayer_reuses_session_when_library_differs(self):
        # Banco isolado: este endpoint escreve em `imports`/`videos`
        # (ensure_video_for_jwplayer_id) e não pode alterar a
        # execução (run_id) "atual" do banco real usada por outros
        # testes de exportação.
        path = Path("tests/.test_switch_library_analyze.db")
        path.unlink(missing_ok=True)

        try:
            with patch(
                "app.DATABASE",
                Database(path),
            ), patch.object(
                JW_SESSION,
                "status",
                return_value={
                    "state": "connected",
                    "property_id": "FvJr6FNj",
                },
            ), patch.object(
                JW_SESSION,
                "switch_property",
                return_value={
                    "state": "connected",
                    "connected": True,
                    "property_id": "XdfUPSCL",
                },
            ) as mock_switch, patch(
                "app.check_publish_dates",
                return_value={"will_be_analyzed": 1},
            ), patch(
                "app.enqueue_jobs",
                return_value=[],
            ) as mock_enqueue:
                response = self.client.post(
                    "/api/analyze-jwplayer",
                    json={
                        "jwplayer_id": "SwiTest1",
                        "library": "VIDEOSSANAR",
                        "property_id": "XdfUPSCL",
                        "provider": "Gemini",
                        "model": "gemini-flash-latest",
                    },
                )

            self.assertEqual(response.status_code, 200)
            mock_switch.assert_called_once_with("XdfUPSCL")
            mock_enqueue.assert_called_once()
        finally:
            path.unlink(missing_ok=True)
            for wal in Path("tests").glob(
                ".test_switch_library_analyze.db-*"
            ):
                wal.unlink(missing_ok=True)

    def test_session_interruption_is_not_a_media_error(self):
        self.assertTrue(is_session_interruption(JWSessionError("Conecte uma sessão JW Player antes de acessar a mídia.")))
        self.assertTrue(is_session_interruption(JWSessionError("O navegador JW Player foi fechado.")))
        self.assertFalse(is_session_interruption(JWSessionError("Nenhuma fonte HLS ou MP4 foi capturada.")))
        self.assertFalse(is_session_interruption(AIError("Ollama: modelo não encontrado")))

    def _read_csv(self, response):
        text = response.content.decode("utf-8-sig")
        rows = list(csv.reader(io.StringIO(text)))
        return rows[0], rows[1:]

    def test_export_csv_only_includes_completed_lessons(self):
        portfolio = _dedupe_by_jwplayer_id(DATABASE.list_portfolio())
        expected = [row for row in portfolio if row.get("status") == "Concluído"]

        response = self.client.get("/api/export.csv")
        self.assertEqual(response.status_code, 200)

        header, data_rows = self._read_csv(response)
        self.assertEqual(len(data_rows), len(expected))

    def test_export_csv_has_no_duplicate_jwplayer_ids(self):
        response = self.client.get("/api/export.csv")
        header, data_rows = self._read_csv(response)

        jwplayer_id_index = header.index("JWPlayer ID")
        ids = [row[jwplayer_id_index] for row in data_rows]

        self.assertEqual(len(ids), len(set(ids)))

    def test_videos_endpoint_has_no_duplicate_jwplayer_ids(self):
        response = self.client.get("/api/videos")
        self.assertEqual(response.status_code, 200)

        items = response.json()["items"]
        ids = [item["jwplayer_id"] for item in items]

        self.assertEqual(len(ids), len(set(ids)))

    def test_dedupe_by_jwplayer_id_keeps_first_occurrence(self):
        rows = [
            {"jwplayer_id": "abc123", "lesson_name": "Aula 1"},
            {"jwplayer_id": "abc123", "lesson_name": "Aula 1 duplicada"},
            {"jwplayer_id": "xyz456", "lesson_name": "Aula 2"},
            {"jwplayer_id": "xyz456", "lesson_name": "Aula 2 duplicada"},
            {"jwplayer_id": "xyz456", "lesson_name": "Aula 2 triplicada"},
        ]

        deduped = _dedupe_by_jwplayer_id(rows)

        self.assertEqual(
            [(row["jwplayer_id"], row["lesson_name"]) for row in deduped],
            [("abc123", "Aula 1"), ("xyz456", "Aula 2")],
        )

    def test_export_csv_excludes_forbidden_columns(self):
        response = self.client.get("/api/export.csv")
        header, _ = self._read_csv(response)
        self.assertFalse(FORBIDDEN_CSV_COLUMNS.intersection(header))

    def test_export_csv_filters_by_publish_year(self):
        portfolio = _dedupe_by_jwplayer_id(DATABASE.list_portfolio())
        completed = [row for row in portfolio if row.get("status") == "Concluído"]

        years = {
            _publish_date_year(row.get("publish_date"))
            for row in completed
        }
        years.discard(None)
        self.assertTrue(years, "esperava ao menos um ano entre as aulas concluídas")

        year = sorted(years)[0]
        expected = [
            row
            for row in completed
            if _publish_date_year(row.get("publish_date")) == year
        ]

        response = self.client.get("/api/export.csv", params={"year": str(year)})
        self.assertEqual(response.status_code, 200)

        _, data_rows = self._read_csv(response)
        self.assertEqual(len(data_rows), len(expected))

    def test_export_csv_rejects_invalid_year(self):
        response = self.client.get("/api/export.csv", params={"year": "not-a-year"})
        self.assertEqual(response.status_code, 422)

    def test_publish_date_year_handles_timezone_and_z_suffix(self):
        self.assertEqual(_publish_date_year("2025-05-13T11:00:18+00:00"), 2025)
        self.assertEqual(_publish_date_year("2025-05-13T11:00:18Z"), 2025)
        self.assertEqual(_publish_date_year("2025-05-13T23:59:00-03:00"), 2025)
        self.assertEqual(_publish_date_year("2025-05-13"), 2025)
        self.assertIsNone(_publish_date_year(None))
        self.assertIsNone(_publish_date_year(""))
        self.assertIsNone(_publish_date_year("not-a-date"))

    def test_publish_date_year_ignores_a_future_year_it_does_not_have(self):
        # A lógica de extração do ano é agnóstica ao ano corrente —
        # qualquer ano presente no publish_date é reconhecido,
        # inclusive anos futuros (ex.: 2027), sem regra especial.
        self.assertEqual(_publish_date_year("2027-01-15T09:00:00Z"), 2027)

    def test_export_csv_includes_publish_year_column(self):
        portfolio = _dedupe_by_jwplayer_id(DATABASE.list_portfolio())
        completed = [row for row in portfolio if row.get("status") == "Concluído"]

        response = self.client.get("/api/export.csv")
        header, data_rows = self._read_csv(response)

        self.assertIn("Ano de publicação", header)

        year_index = header.index("Ano de publicação")
        exported_years = [row[year_index] for row in data_rows]
        expected_years = [
            (
                str(_publish_date_year(row.get("publish_date")))
                if _publish_date_year(row.get("publish_date")) is not None
                else ""
            )
            for row in completed
        ]
        self.assertEqual(exported_years, expected_years)

    def test_rows_for_export_is_the_single_source_of_year_logic(self):
        # /api/videos (filtro da tela), /api/export.csv e
        # /api/export.xlsx precisam concordar sobre o que é
        # "o ano de publicação" de uma linha — todos usam
        # _publish_date_year() através de _rows_for_export().
        rows = _rows_for_export()
        for row in rows:
            self.assertEqual(
                row["_export_year"],
                _publish_date_year(row.get("publish_date")),
            )

    def _load_xlsx(self, response):
        return load_workbook(io.BytesIO(response.content))

    def test_export_xlsx_has_videos_and_summary_sheets(self):
        response = self.client.get("/api/export.xlsx")
        self.assertEqual(response.status_code, 200)

        workbook = self._load_xlsx(response)
        self.assertEqual(workbook.sheetnames, ["Vídeos", "Resumo por ano"])

    def test_export_xlsx_videos_sheet_matches_export_rules(self):
        expected_rows = _rows_for_export()

        response = self.client.get("/api/export.xlsx")
        workbook = self._load_xlsx(response)
        videos_sheet = workbook["Vídeos"]

        header = [cell.value for cell in videos_sheet[1]]
        self.assertIn("Ano de publicação", header)
        self.assertNotIn("Publish date", header)
        for forbidden in FORBIDDEN_CSV_COLUMNS:
            self.assertNotIn(forbidden, header)

        data_rows = list(
            videos_sheet.iter_rows(min_row=2, values_only=True)
        )
        self.assertEqual(len(data_rows), len(expected_rows))

        jwplayer_id_index = header.index("JWPlayer ID")
        ids = [row[jwplayer_id_index] for row in data_rows]
        self.assertEqual(len(ids), len(set(ids)))

        self.assertTrue(videos_sheet.auto_filter.ref)

    def test_export_xlsx_summary_matches_videos_sheet(self):
        response = self.client.get("/api/export.xlsx")
        workbook = self._load_xlsx(response)

        videos_sheet = workbook["Vídeos"]
        header = [cell.value for cell in videos_sheet[1]]
        year_index = header.index("Ano de publicação")

        totals_from_videos = {}
        for row in videos_sheet.iter_rows(min_row=2, values_only=True):
            year = row[year_index]
            label = str(year) if year is not None else "Não informado"
            totals_from_videos[label] = totals_from_videos.get(label, 0) + 1

        summary_sheet = workbook["Resumo por ano"]
        self.assertEqual(
            [cell.value for cell in summary_sheet[1]],
            ["Ano de publicação", "Total de vídeos"],
        )

        totals_from_summary = {
            str(label): total
            for label, total in summary_sheet.iter_rows(
                min_row=2, values_only=True
            )
        }

        self.assertEqual(totals_from_summary, totals_from_videos)
        self.assertEqual(
            sum(totals_from_summary.values()),
            len(_rows_for_export()),
        )

    # TESTE 1 — a exportação possui exatamente 13 colunas.
    def test_export_has_exactly_thirteen_columns(self):
        self.assertEqual(len(EXPORT_COLUMNS), 13)

        response = self.client.get("/api/export.csv")
        header, _ = self._read_csv(response)
        self.assertEqual(len(header), 13)

    # TESTE 2 — ordem exata das colunas.
    def test_export_column_order_matches_specification(self):
        expected_order = [
            "Nome da aula",
            "Modelo de aula",
            "Professor",
            "Resumo do conteúdo",
            "JWPlayer ID",
            "Status",
            "Validação Manual",
            "Confiança",
            "Palavras-chave",
            "Ano de publicação",
            "Macrotema",
            "Microtema",
            "Nanotema",
        ]

        self.assertEqual(
            [label for _, label in EXPORT_COLUMNS], expected_order
        )

        response = self.client.get("/api/export.csv")
        header, _ = self._read_csv(response)
        self.assertEqual(header, expected_order)

    # TESTE 3 — a coluna "Link da aula" não existe (nem no CSV, nem
    # no XLSX).
    def test_export_does_not_have_a_separate_link_column(self):
        csv_response = self.client.get("/api/export.csv")
        csv_header, _ = self._read_csv(csv_response)
        self.assertNotIn("Link da aula", csv_header)

        xlsx_response = self.client.get("/api/export.xlsx")
        workbook = self._load_xlsx(xlsx_response)
        xlsx_header = [cell.value for cell in workbook["Vídeos"][1]]
        self.assertNotIn("Link da aula", xlsx_header)

    # TESTE 4 e 5 — a coluna "Modelo de aula" existe e reflete
    # exatamente o valor já persistido em final_category, sem
    # recalcular nada na exportação.
    def test_export_includes_modelo_de_aula_from_final_category(self):
        portfolio = _dedupe_by_jwplayer_id(DATABASE.list_portfolio())
        completed = [
            row for row in portfolio if row.get("status") == "Concluído"
        ]
        expected = [row.get("final_category") or "" for row in completed]

        response = self.client.get("/api/export.csv")
        header, data_rows = self._read_csv(response)

        self.assertIn("Modelo de aula", header)
        index = header.index("Modelo de aula")
        self.assertEqual([row[index] for row in data_rows], expected)

    def test_export_csv_defaults_missing_topics_to_nao_identificado(self):
        response = self.client.get("/api/export.csv")
        header, data_rows = self._read_csv(response)

        for field in ("Macrotema", "Microtema", "Nanotema"):
            index = header.index(field)
            for row in data_rows:
                self.assertTrue(row[index])

    # TESTE 6, 7 e 8 — no Excel, a própria célula "JWPlayer ID"
    # possui hyperlink, apontando para a mesma URL usada pelo site
    # (build_jwplayer_media_url), e o texto exibido continua sendo
    # somente o JWPlayer ID.
    def test_export_xlsx_jwplayer_id_cell_is_the_hyperlink(self):
        response = self.client.get("/api/export.xlsx")
        workbook = self._load_xlsx(response)
        videos_sheet = workbook["Vídeos"]

        header = [cell.value for cell in videos_sheet[1]]
        jwplayer_col = header.index("JWPlayer ID") + 1
        library = get_current_library()

        checked_any = False
        for row in videos_sheet.iter_rows(min_row=2):
            jwplayer_cell = row[jwplayer_col - 1]
            jwplayer_id = jwplayer_cell.value

            expected_url = build_jwplayer_media_url(
                jwplayer_id, library["property_id"]
            )

            if expected_url:
                self.assertEqual(
                    jwplayer_cell.hyperlink.target, expected_url
                )
                self.assertEqual(jwplayer_cell.value, jwplayer_id)
                self.assertFalse(
                    str(jwplayer_cell.value).startswith("http")
                )
                checked_any = True

        self.assertTrue(checked_any, "esperava ao menos um link clicável")

    # TESTE 10 — Macrotema, Microtema e Nanotema continuam presentes.
    def test_export_includes_topic_classification_columns(self):
        response = self.client.get("/api/export.csv")
        header, _ = self._read_csv(response)
        for column in ("Macrotema", "Microtema", "Nanotema"):
            self.assertIn(column, header)


if __name__ == "__main__":
    unittest.main()
